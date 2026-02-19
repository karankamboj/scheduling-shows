import re
from datetime import datetime, date
import openpyxl
import pandas as pd



pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)

FILE_PATH = "/Users/karankamboj/Documents/Work/Scheduling/data.xlsx"

MOD_ACT_RE = re.compile(r"M\d+\s*A\d+", re.IGNORECASE)

def to_iso(d):
    """Convert Excel datetime/date to ISO string; pass through None."""
    if d is None or d == "":
        return None
    if isinstance(d, datetime):
        return d.date().isoformat()
    if isinstance(d, date):
        return d.isoformat()
    return str(d).strip()

def parse_subject_open_close(ws, max_rows=500):
    """
    Extract:
      Subject (e.g., 'Bio 181 Mod/Act'),
      Mod/Act (e.g., 'M1 A1'),
      VR Show Open Date,
      VR Show Close Date*
    from 'Highlevel and Academi*' sheets.
    """
    rows = []
    current_subject = None

    for r in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=6, values_only=True):
        c1, c2, c3, c4, *_ = (list(r) + [None] * 6)[:6]

        # Subject header row (e.g., "Bio 181 Mod/Act")
        if isinstance(c1, str) and "mod/act" in c1.lower():
            cleaned = c1.replace("\n", " ")
            
            # Remove Mod/Act (case-insensitive)
            cleaned = re.sub(r"mod/act", "", cleaned, flags=re.IGNORECASE)
            
            # Normalize spaces
            current_subject = " ".join(cleaned.split())
            
            continue

        # Data row: Mod/Act + dates
        if current_subject and isinstance(c1, str) and MOD_ACT_RE.match(c1):
            mod_act = c1.strip()
            open_date = to_iso(c3)
            close_date = to_iso(c4)

            # only keep rows that actually have dates
            if open_date or close_date:
                rows.append(
                    {
                        # "sheet": ws.title,
                        "subject": current_subject,
                        "mod_act": mod_act,
                        "vr_show_open_date": open_date,
                        "vr_show_close_date": close_date,
                    }
                )

    return pd.DataFrame(rows)

def parse_course_caps(ws, max_rows=500, max_cols=30):
    """
    Extract a table that contains headers like:
      'Course' and '@ Cap'
    """
    header_row_idx = None
    header = None

    # Find the header row
    for ridx, r in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True),
        start=1,
    ):
        values = [v.strip() if isinstance(v, str) else v for v in r]
        if any(v == "Course" for v in values) and any(isinstance(v, str) and "cap" in v.lower() for v in values):
            header_row_idx = ridx
            header = values
            break

    if header_row_idx is None:
        return pd.DataFrame(columns=["sheet", "course", "cap"])

    course_col = header.index("Course") + 1
    cap_col = next(i for i, v in enumerate(header) if isinstance(v, str) and "cap" in v.lower()) + 1

    rows = []
    for r in ws.iter_rows(
        min_row=header_row_idx + 1,
        max_row=header_row_idx + 200,  # enough to grab the table
        min_col=1,
        max_col=max_cols,
        values_only=True,
    ):
        course = r[course_col - 1]
        cap = r[cap_col - 1]

        if course is None or (isinstance(course, str) and course.strip() == ""):
            # stop when the course column becomes empty (end of table)
            if rows:
                break
            continue

        rows.append(
            {
                "sheet": ws.title,
                "course": str(course).strip(),
                "cap": cap,
            }
        )

    return pd.DataFrame(rows)

def parse_holidays(ws, max_rows=2000):
    """
    Extract holiday dates listed under a 'HOLIDAYS' label (usually near bottom).
    """
    holidays_row = None
    holidays_col = None

    # Find the 'HOLIDAYS' cell
    for ridx, r in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        for cidx, v in enumerate(r, start=1):
            if isinstance(v, str) and v.strip().lower() == "holidays":
                holidays_row = ridx
                holidays_col = cidx
                break
        if holidays_row:
            break

    if holidays_row is None:
        return pd.DataFrame(columns=["sheet", "holiday_date"])

    # Read dates below it until blank
    rows = []
    for ridx in range(holidays_row + 1, holidays_row + 200):
        v = ws.cell(row=ridx, column=holidays_col).value
        if v is None or v == "":
            if rows:
                break
            continue
        rows.append({"sheet": ws.title, "holiday_date": to_iso(v)})

    return pd.DataFrame(rows)

def main():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)

    # Typically these are the sheets that contain the requested sections:
    target_sheets = [name for name in wb.sheetnames if "Highlevel and Academi" in name]

    all_subjects = []
    all_course_caps = []
    all_holidays = []

    for name in target_sheets:
        ws = wb[name]
        all_subjects.append(parse_subject_open_close(ws))
        all_course_caps.append(parse_course_caps(ws))
        all_holidays.append(parse_holidays(ws))

    subjects_df = pd.concat(all_subjects, ignore_index=True) if all_subjects else pd.DataFrame()
    caps_df = pd.concat(all_course_caps, ignore_index=True) if all_course_caps else pd.DataFrame()
    holidays_df = pd.concat(all_holidays, ignore_index=True) if all_holidays else pd.DataFrame()

    print("\n=== Subjects: VR Show Open/Close Dates ===")
    print(subjects_df)

    print("\n=== Course with @Cap ===")
    print(caps_df)

    print("\n=== Holidays (bottom) ===")
    print(holidays_df)

    # Optional: save outputs
    subjects_df.to_csv("subjects_vr_open_close.csv", index=False)
    caps_df.to_csv("course_caps.csv", index=False)
    holidays_df.to_csv("holidays.csv", index=False)

if __name__ == "__main__":
    main()
